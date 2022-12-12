# -*- coding: utf-8 -*-
"""
Created on Fri May 27 18:41:00 2022

"""
from sklearn import linear_model
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_squared_error
import pandas as pd
import numpy as np
from scipy.linalg import svd
from openpyxl.workbook import Workbook
from openpyxl import load_workbook


        
        
    
        
def start():
    #ap=APP()
    TestFile='test.csv'
    dfTest = pd.read_csv(TestFile)
    
    TrainFile='user_artist.csv'
    dfTrain = pd.read_csv(TrainFile)
    
    users=list(set(dfTrain['userID']))
    Nu=len(users)  #number of users
    Artists=list(set(dfTrain['artistID']))
    #ArtistsList=set(dfTrain['artistID'])
    
    
    # preparing the Training set
    #***************************************************************
    #for each user , preaper feteurs vector at length of number of artist with the number of hearings(weight).
    
    A=np.zeros([len(dfTrain),len(users)+len(Artists)])


    C=np.log10(dfTrain['weight'])
    for i in range(len(dfTrain)):
        if not(i%100):
            print(i)
        R_ind=users.index(dfTrain['userID'][i])
        C_ind=Artists.index(dfTrain['artistID'][i])    
        A[i,R_ind]=1
        A[i,Nu+C_ind]=1    

    return(A,users,Artists,C,dfTrain,dfTest)

   



def lossFun(R,P,Q,g):  #SVD
    
    #P=np.reshape(alpha[0:n], [u,rank])
    #Q=np.reshape(alpha[n:m],[rank,i])
    #G=np.diag(alpha[m:])
    G=np.diag(g)
    R_i= R-np.matmul(P,np.matmul(G,Q))
    L=np.sum((R_i)**2)
    
    return (L , R_i)
    

def Latent(FV):  #SVD
    # perform SVD
    rank =(np.linalg.matrix_rank(FV));
    #rank=round(rank/2)
    #rank=100
    U, singular, V_transpose = svd(FV,full_matrices=False)
    P=U[:,0:rank]
    Q=V_transpose[0:rank,:]
    g=singular[0:rank]
    return (P , Q ,g,rank)


def LossTest(N,RegOpt,dfTrain,indx_test,users,Artist,d,R_tilda):
    bu=RegOpt.coef_[0:len(users)]
    bi=RegOpt.coef_[len(users):] 
    L=0
    rdif=0
    for i in indx_test: 
        userID=dfTrain['userID'][i]
        artistID=dfTrain['artistID'][i]
        ind=np.argsort(np.abs(d[:,np.where(Artist==artistID)]))
        neighbo=0
        
        for k in range(1,N):
            neighbo+=d[ind[-k],np.where(Artist==artistID)]*R_tilda[np.where(users==userID),ind[-k]]
        if np.sum(np.abs(d[ind[-N:],np.where(Artist==artistID)]))>0:
            neighbo=neighbo/np.sum(np.abs(d[ind[-N:],np.where(Artist==artistID)]))
        else:
            neighbo=0
        
        rui=ravg+bu[np.where(users==userID)]+bi[np.where(Artist==artistID)]
        ruiN=ravg+bu[np.where(users==userID)]+bi[np.where(Artist==artistID)]+neighbo
        
        r=np.log10(dfTrain['weight'][i])
        if N>0:
            L+=(ruiN-r)**2/len(indx_test)
        else:
            L+=(rui-r)**2/len(indx_test)
        if rui!=ruiN:
            rdif+=(rui-ruiN)**2
            print(rdif)
        print('predict %f :predict neighbo %f actual %f  Loos %f'%(rui,ruiN,r,L))
    return (L,rdif)
    
def Predict_Bias_Only(RegOpt,dfTrain,indx_test,users,Artist,ravg):
    bu=RegOpt.coef_[0:len(users)]
    bi=RegOpt.coef_[len(users):]
    j=0
    Y=np.zeros(len(indx_test))
    for i in indx_test:
        userID=dfTrain['userID'][i]
        artistID=dfTrain['artistID'][i]
        rui=ravg+bu[np.where(users==userID)]+bi[np.where(Artist==artistID)]
        Y[j]=rui
        j+=1
    return(Y)

def Predict_Bias_and_Neighbours(RegOpt,dfTest,indx_test,users,Artists,ravg,d,R_tilda,N):
    bu=RegOpt.coef_[0:len(users)]
    bi=RegOpt.coef_[len(users):]
    j=0
    Y=np.zeros(len(indx_test))
    for i in indx_test:
        userID=dfTest['userID'][i]
        artistID=dfTest['artistID'][i]
        ind=np.argsort(np.abs(d[:,np.where(Artists==artistID)]))
        neighbo=0
        for k in range(1,N):
            neighbo+=d[ind[-k],np.where(Artists==artistID)]*R_tilda[np.where(users==userID),ind[-k]]
        if np.sum(np.abs(d[ind[-N:],np.where(Artists==artistID)]))>0:
            neighbo=neighbo/np.sum(np.abs(d[ind[-N:],np.where(Artists==artistID)]))
        else:
            neighbo=0
        ruiN=ravg+bu[np.where(users==userID)]+bi[np.where(Artists==artistID)]+neighbo
        if ruiN.size>0:
            Y[j]=ruiN
        else:
            Y[j]=ravg #'nan'
        j+=1
    return(Y)
    
#Load tha data and set the following array
##########################################3 
A,users,Artists,C,dfTrain,dfTest=start()        
##########################################     
#split the dataset into training (70%) and testing (30%) sets
n=len(dfTrain)
ind=list(range(n))
indx_train,indx_test,indy_train,indy_test= train_test_split(ind,ind,test_size=0.3,random_state=0)

#X_train,X_test,y_train,y_test= train_test_split(X,Y,test_size=0.3,random_state=0) 
#X_train=X[indx_train]
X_train=np.matmul((A[indx_train]).T,A[indx_train])

R=np.zeros([len(users),len(Artists)])
ravg=0
for i in indx_train:
    userID=dfTrain['userID'][i]
    artistID=dfTrain['artistID'][i]
    r=np.log10(dfTrain['weight'][i])
    R[np.where(users==userID),np.where(Artists==artistID)]=r
    ravg+=r/len(indx_train)
    
y_train=np.matmul((A[indx_train]).T,C[indx_train]-ravg)
y_test=C[indx_test]

#step 2 : training the Model "Baseline predictors" + "Regularization"
Temp=np.inf
for Lambda in np.linspace(10,20,11):
    print(Lambda)
    
    reg = linear_model.Ridge(alpha=Lambda)
    #reg = linear_model.Ridge()
    RegM=reg.fit(X_train,y_train)
    #test
    
    y_predictTest=Predict_Bias_Only(RegM,dfTrain,indx_test,users,Artists,ravg)
    y_predictTrain=Predict_Bias_Only(RegM,dfTrain,indx_train,users,Artists,ravg)
    #MESTest=mean_squared_error(y_test, y_predictTest, squared = True)
    #MESTest=LossTest(0,RegM,dfTrain,indx_test,users,Artists,0,0)
    MESTest=mean_squared_error(y_test, y_predictTest, squared = True)
    #MESTrain=mean_squared_error(C[indx_train], y_predictTrain, squared = True)
    #print('Train loos %f'%MESTrain)
    print('Test loos %f'%MESTest)
    if MESTest>Temp:
        break
    else:
        Temp=MESTest
        RegOpt=RegM  #the Optimum model
        
#******************************************************************************        
#Step 3: prediction with  Neighborhood Model
#calculating of Artists  similarity metric 𝑑 

R_hat=np.zeros([len(users),len(Artists)])
bu=RegOpt.coef_[0:len(users)]
bi=RegOpt.coef_[len(users):]
for i in indx_train:
    userID=dfTrain['userID'][i]
    artistID=dfTrain['artistID'][i] 
    rui=ravg+bu[np.where(users==userID)]+bi[np.where(Artists==artistID)]
    r=np.log10(dfTrain['weight'][i])
    R_hat[np.where(users==userID),np.where(Artists==artistID)]=rui
        
        
R_tilda=R-R_hat
D=np.matmul(R_tilda.T,R_tilda)
d=np.zeros([len(Artists),len(Artists)])
for i in range(len(Artists)):
    if not (i%500):
        print(i)
    for j in range(i):
            d[i,j]=D[i,j]/(np.sum((R_tilda[:,i])**2)*np.sum((R_tilda[:,j])**2))**(0.5)
            d[j,i]=D[i,j]
            
for i in range(len(d)):
    d[:,i]=np.nan_to_num(d[:,i])
#***************************************************************************************************    
#step 4 : calculate the Loss function with Neighbourhood method
N=100 # number of Neighbours
L,rdif=LossTest(N,RegOpt,dfTrain,indx_test,users,Artists,d,R_tilda) 

print('loos (𝑅𝑀𝑆𝐸) on the test dataset  = %f'%L)   
#***************************************************************************************************

#Last part : fill the predicted weigth in the test file
#we generate new test.xlsx file for writing the predicted results 
N=100
TestFile='test.xlsx'
Test=Workbook()
Test=load_workbook(TestFile)
page = Test.active
for i in range(len(dfTest)):
    ruiN =Predict_Bias_and_Neighbours(RegOpt,dfTest,[i],users,Artists,ravg,d,R_tilda,N)
    W=10**ruiN[0]    
    page.cell(i+2,3,W)    
Test.save(TestFile) 
            
